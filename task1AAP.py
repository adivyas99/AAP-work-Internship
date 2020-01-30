#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 21 07:44:38 2019

@author: adityavyas
"""

import pandas as pd
df = pd.read_excel('loan.xlsx')

col = df.columns.tolist()
print(col)
for i in col:
    print(i)

if 
df.sheet_names
df.
from openpyxl import load_workbook

wb = load_workbook('loan.xlsx', read_only=True)   # open an Excel file and return a workbook

if 'Risk_Score' in df.columns:
    print("hi")
    
av = wb.get_sheet_by_name


from openpyxl import load_workbook

wb = load_workbook(file_workbook, read_only=True)   # open an Excel file and return a workbook

if 'sheet1' in wb.sheetnames:
    print('sheet1 exists')
